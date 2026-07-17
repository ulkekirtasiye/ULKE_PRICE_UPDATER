from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelManager:

    def __init__(self):

        self.yesil = PatternFill(
            fill_type="solid",
            start_color="90EE90",
            end_color="90EE90"
        )

        self.sari = PatternFill(
            fill_type="solid",
            start_color="FFF59D",
            end_color="FFF59D"
        )

        self.kirmizi = PatternFill(
            fill_type="solid",
            start_color="FF9999",
            end_color="FF9999"
        )

    def ac(self, dosya):

        wb = load_workbook(dosya)
        ws = wb.active

        return wb, ws

    def kaydet(self, wb, dosya):

        wb.save(dosya)

    def basliklari_bul(self, ws):

        basliklar = {}

        for hucre in ws[1]:

            if hucre.value is None:
                continue

            basliklar[str(hucre.value).strip().upper()] = hucre.column

        return basliklar