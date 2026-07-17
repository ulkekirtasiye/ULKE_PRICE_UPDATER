from openpyxl.styles import PatternFill


class Updater:

    def __init__(self):

        self.green = PatternFill(
            fill_type="solid",
            start_color="90EE90",
            end_color="90EE90"
        )

        self.yellow = PatternFill(
            fill_type="solid",
            start_color="FFF59D",
            end_color="FFF59D"
        )

    def update_prices(
        self,
        supplier_ws,
        akinsoft_ws,
        supplier_columns,
        akinsoft_columns
    ):

        supplier_products = {}

        # TEDARİKÇİ LİSTESİNİ SÖZLÜĞE AKTAR

        for row in range(2, supplier_ws.max_row + 1):

            barkod = supplier_ws.cell(
                row,
                supplier_columns["barkod"]
            ).value

            fiyat = supplier_ws.cell(
                row,
                supplier_columns["fiyat"]
            ).value

            if barkod is None:
                continue

            supplier_products[str(barkod).strip()] = fiyat

        updated = 0
        not_found = 0

        # AKINSOFT GÜNCELLE

        for row in range(2, akinsoft_ws.max_row + 1):

            barkod = akinsoft_ws.cell(
                row,
                akinsoft_columns["barkod"]
            ).value

            fiyat_hucre = akinsoft_ws.cell(
                row,
                akinsoft_columns["alis"]
            )

            if barkod is None:

                fiyat_hucre.fill = self.yellow
                not_found += 1
                continue

            barkod = str(barkod).strip()

            if barkod in supplier_products:

                fiyat_hucre.value = supplier_products[barkod]
                fiyat_hucre.fill = self.green

                updated += 1

            else:

                fiyat_hucre.fill = self.yellow
                not_found += 1

        return updated, not_found